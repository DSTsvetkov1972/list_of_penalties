from PySide6 import QtCore
# from colorama import Fore
from my_functions.main_window import load_file_sheet_name
import global_vars
import os
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import pandas as pd
from datetime import datetime
from win32com.shell import shell, shellcon


class LoadInDWHThread(QtCore.QThread):
    def __init__(self, parent=None):
        QtCore.QThread.__init__(self, parent)

    def run(self):
        try:
            load_file_sheet_name()
            global_vars.ui.footer_label.setStyleSheet('color: blue')
            global_vars.ui.footer_label.setText("Подготавливаем данные к загрузке в DWH...")

            global_vars.df_to_insert = global_vars.df_to_insert.map(str)

            global_vars.equipments_docs_df = global_vars.equipments_docs_df.map(str)

            global_vars.ui.footer_label.setStyleSheet('color: blue')
            global_vars.ui.footer_label.setText("Шаг 1 из 4. Подтягиваем информацию из БД...")

            res_df = pd.merge(global_vars.df_to_insert, global_vars.equipments_docs_df, how='left', on=['sent_number_equipment', 'order_id'])

            global_vars.ui.footer_label.setStyleSheet('color: blue')
            global_vars.ui.footer_label.setText(f"Шаг 2 из 4. Записываем результат в файл Эксель {len(res_df)} строк...")
            res_file_name = os.path.join(shell.SHGetKnownFolderPath(shellcon.FOLDERID_Downloads),
                                         f'{global_vars.prog_name}_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx')

            res_df.to_excel(res_file_name, index=False)

            global_vars.ui.footer_label.setStyleSheet('color: blue')
            global_vars.ui.footer_label.setText("Шаг 3 из 4. Подгоняем ширины столбцов...")

            wb = load_workbook(res_file_name)
            ws = wb.active

            for n, column in enumerate(res_df.columns, 1):
                # print(n, column)
                ws.column_dimensions[get_column_letter(n)].width = len(str(column))*1.1+1

        except Exception as e:
            global_vars.ui.footer_label.setText(f"{e}. Строк {len(res_df)}")

        global_vars.ui.footer_label.setStyleSheet('color: blue')
        global_vars.ui.footer_label.setText("Шаг 4 из 4. Замораживаем строку заголовока...")
        ws.freeze_panes = ws.cell(column=1, row=2)

        wb.save(res_file_name)

        global_vars.ui.footer_label.setStyleSheet('color: green')
        global_vars.ui.footer_label.setText(f"Результат сохранен в загрузках. Имя файла: {res_file_name}")

        os.startfile(res_file_name)

    def on_started(self):
        global_vars.ui.verticalLayoutWidgetButtons_1.setEnabled(False)
        global_vars.ui.verticalLayoutWidgetButtons_2.setEnabled(False)
        global_vars.ui.comboSheets.setEnabled(False)
        global_vars.ui.footer_text.setVisible(False)

    def on_finished(self):
        global_vars.ui.verticalLayoutWidgetButtons_1.setEnabled(True)
        global_vars.ui.verticalLayoutWidgetButtons_2.setEnabled(True)
        global_vars.ui.comboSheets.setEnabled(True)
