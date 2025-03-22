from PySide6 import QtCore, QtWidgets
from colorama import Fore
from my_functions.dwh import execute_sql_click, insert_from_df, get_df_of_click
from my_functions.main_window import translit, load_file_sheet_name
import global_vars
import os
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
import pandas as pd
from datetime import datetime
from win32com.shell import shell, shellcon

class LoadInDWHThread(QtCore.QThread):
    def __init__ (self, parent=None):
        QtCore.QThread.__init__(self, parent)


    def run(self):
        load_file_sheet_name()
        global_vars.ui.footer_label.setStyleSheet('color: blue')
        global_vars.ui.footer_label.setText(f"Подготавливаем данные к загрузке в DWH...") 

   
            

        #df_to_Insert = preprocessing(global_vars.df, global_vars.header_row, global_vars.horizontal_headers, global_vars.column_formats, rows_number_column_name)

        global_vars.ui.footer_label.setStyleSheet('color: blue')        
        global_vars.ui.footer_label.setText(f"Создаём таблицу {global_vars.dwh_table_name} в DWH...") 

        global_vars.df_to_insert = global_vars.df_to_insert.map(str)
        global_vars.equipments_docs_df = global_vars.equipments_docs_df.map(str)
        print(global_vars.df_to_insert.info())
        print(global_vars.equipments_docs_df.info())        

        global_vars.ui.footer_label.setStyleSheet('color: blue')            
        global_vars.ui.footer_label.setText("Шаг 1 из 3. Подтягиваем информацию из БД...")

        res_df = pd.merge(global_vars.df_to_insert, global_vars.equipments_docs_df, how='left', on=['sent_number_equipment', 'order_id'])



        global_vars.ui.footer_label.setStyleSheet('color: blue')            
        global_vars.ui.footer_label.setText("Шаг 2 из 3. Записываем результат в файл Эксель...")
        res_file_name = os.path.join(shell.SHGetKnownFolderPath(shellcon.FOLDERID_Downloads),f'{global_vars.prog_name}_{datetime.now().strftime("%Y-%m-%d_%H-%M-%S")}.xlsx')

        res_df.to_excel(res_file_name, index=False)

        global_vars.ui.footer_label.setStyleSheet('color: blue')            
        global_vars.ui.footer_label.setText("Шаг 3 из 3. Подгоняем ширины столбцов...")

        wb = load_workbook(res_file_name)
        ws = wb.active

        for n, column in enumerate(res_df.columns, 1):
            # print(n, column)
            ws.column_dimensions[get_column_letter(n)].width = len(str(column))*1.1+1
        
        ws.freeze_panes = ws.cell(column=1, row=2)

        wb.save(res_file_name)



        global_vars.ui.footer_label.setStyleSheet('color: green')            
        global_vars.ui.footer_label.setText(f"Результат сохранен в загрузках. Имя файла: {res_file_name}")

        os.startfile(res_file_name)


        


    def on_started(self): # Вызывается при запуске потока
        global_vars.ui.verticalLayoutWidgetButtons_1.setEnabled(False)        
        global_vars.ui.verticalLayoutWidgetButtons_2.setEnabled(False)
        global_vars.ui.comboSheets.setEnabled(False) 
        global_vars.ui.footer_text.setVisible(False)          
        


    def on_finished(self): # Вызывается при завершении потока
        global_vars.ui.verticalLayoutWidgetButtons_1.setEnabled(True)
        global_vars.ui.verticalLayoutWidgetButtons_2.setEnabled(True)
        global_vars.ui.comboSheets.setEnabled(True)
